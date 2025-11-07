import os
import logging
from typing import List, Optional, Tuple, Any
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from utils.file_treatment import (
    _generate_unique_folder,
    _generate_filename,
    _public_url,
)

def add_auto_sized_review_comment(cell, text: str, author: str = "AI Reviewer") -> None:
    """
    Adds a note to an Excel cell, adjusting the width and height so that all the text is visible.
    """
    if not text:
        return

    avg_char_width = 7
    px_per_line = 15
    base_width = 200
    max_width = 500
    min_height = 40

    width = min(max_width, base_width + len(text) * 2)
    chars_per_line = max(1, width // avg_char_width)
    lines = 0
    for paragraph in text.split("\n"):
        lines += -(-len(paragraph) // chars_per_line)  # ceil division
    height = max(min_height, lines * px_per_line)

    comment = Comment(text, author)
    comment.width = width
    comment.height = height
    cell.comment = comment


def create_excel(
    data: List[List[Any]],
    filename: Optional[str],
    folder_path: Optional[str] = None,
    title: Optional[str] = None,
    xlsx_template_path: Optional[str] = None,
) -> dict:
    """
    Create an XLSX file using an optional template.

    Returns: {"url": public_url, "path": local_path}
    """
    log = logging.getLogger(__name__)
    log.debug("Creating Excel file with optional template")

    if folder_path is None:
        folder_path = _generate_unique_folder()

    if filename:
        filepath = os.path.join(folder_path, filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        fname = filename
    else:
        filepath, fname = _generate_filename(folder_path, "xlsx")

    # Load template if provided
    if xlsx_template_path:
        try:
            log.debug(f"Loading XLSX template from {xlsx_template_path} ...")
            wb = load_workbook(xlsx_template_path)
            log.debug(f"Template loaded with {len(wb.sheetnames)} sheet(s)")
        except Exception as e:
            log.warning(f"Failed to load XLSX template: {e}")
            wb = Workbook()
    else:
        log.debug("No XLSX template, creating new workbook")
        wb = Workbook()

    ws = wb.active

    from openpyxl.utils import get_column_letter

    if title:
        # apply safe sheet title
        ws.title = "".join(c for c in title if c.isalnum() or c in (" ", "-", "_")).strip()[:31]
        # try replace a cell containing 'title' with actual title
        title_cell_found = False
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "title" in cell.value.lower():
                        cell.value = title
                        try:
                            from openpyxl.styles import Font
                            cell.font = Font(bold=True)
                        except Exception:
                            pass
                        log.debug(
                            f"Title '{title}' replaced at cell {get_column_letter(cell.column)}{cell.row}"
                        )
                        title_cell_found = True
                        break
                if title_cell_found:
                    break
        except Exception:
            pass

    # Start position can be influenced by auto_filter in template
    start_row, start_col = 1, 1
    if ws.auto_filter and ws.auto_filter.ref:
        try:
            from openpyxl.utils import range_boundaries
            start_col, start_row, _, _ = range_boundaries(ws.auto_filter.ref)
        except Exception:
            pass

    if not data:
        wb.save(filepath)
        return {"url": _public_url(folder_path, fname), "path": filepath}

    template_border = ws.cell(start_row, start_col).border
    has_borders = template_border and any(
        [
            template_border.top.style,
            template_border.bottom.style,
            template_border.left.style,
            template_border.right.style,
        ]
    )

    # write grid and optionally apply borders
    rows_count = max(len(data) + 10, 50)
    cols_count = max(len(data[0]) + 5, 20) if data and isinstance(data[0], list) else 20

    for r in range(rows_count):
        for c in range(cols_count):
            cell = ws.cell(row=start_row + r, column=start_col + c)

            if r < len(data) and c < len(data[0]):
                cell.value = data[r][c]
                # bold header row
                if r == 0 and data[r][c]:
                    try:
                        from openpyxl.styles import Font
                        cell.font = Font(bold=True)
                    except Exception:
                        pass
                if has_borders:
                    try:
                        from openpyxl.styles import Border
                        cell.border = Border(
                            top=template_border.top,
                            bottom=template_border.bottom,
                            left=template_border.left,
                            right=template_border.right,
                        )
                    except Exception:
                        pass
            else:
                cell.value = None
                # clear residual styles
                try:
                    if cell.has_style:
                        from openpyxl.styles import Font, PatternFill, Border, Alignment
                        cell.font, cell.fill, cell.border, cell.alignment = (
                            Font(),
                            PatternFill(),
                            Border(),
                            Alignment(),
                        )
                except Exception:
                    pass

    # reset auto_filter to actual used range
    try:
        if data and ws.auto_filter:
            ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + len(data[0]) - 1)}{start_row + len(data) - 1}"
    except Exception:
        pass

    # autosize columns
    try:
        if data:
            for c in range(len(data[0])):
                max_len = 0
                for r in range(len(data)):
                    val = data[r][c]
                    max_len = max(max_len, len(str(val)) if val is not None else 0)
                ws.column_dimensions[get_column_letter(start_col + c)].width = min(max_len + 2, 150)
    except Exception:
        pass

    wb.save(filepath)
    return {"url": _public_url(folder_path, fname), "path": filepath}
